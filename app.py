import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import os
import glob
from datetime import datetime, timedelta
from pymongo import MongoClient

st.set_page_config(
    page_title="Mapa de Vendas",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded"
)

PASSWORD = "SCH@6013"

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

def check_password():
    if not st.session_state.authenticated:
        st.markdown("""
        <style>
        .main {background-color: #f0f2f6;}
        .stTextInput > div > div > input {text-align: center; font-size: 18px;}
        </style>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown("<br><br>", unsafe_allow_html=True)
            st.markdown("## Mapa de Vendas 2026")
            st.markdown("#### Introduza a password para aceder")
            password = st.text_input("Password", type="password", key="pwd_input")
            if st.button("Entrar", use_container_width=True):
                if password == PASSWORD:
                    st.session_state.authenticated = True
                    st.rerun()
                else:
                    st.error("Password incorreta!")
            st.markdown("---")
            st.caption("Dashboard de Vendas Mensais")
        return False
    return True

if not check_password():
    st.stop()

def _load_env_file():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "atlas-credentials.env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, value = line.split("=", 1)
                    os.environ.setdefault(key.strip(), value.strip().strip('"'))

def _get_mongo_uri():
    mongo_uri = os.environ.get("MONGODB_URI", "")
    if not mongo_uri:
        _load_env_file()
        mongo_uri = os.environ.get("MONGODB_URI", "")
    return mongo_uri

def get_db():
    if "mongo_db" in st.session_state:
        return st.session_state.mongo_db
    mongo_uri = _get_mongo_uri()
    if not mongo_uri:
        return None
    try:
        client = MongoClient(mongo_uri, serverSelectionTimeoutMS=5000, connectTimeoutMS=5000, socketTimeoutMS=10000)
        client.admin.command("ping")
        db = client["vendas_dashboard"]
        st.session_state.mongo_db = db
        return db
    except Exception as e:
        st.session_state["mongo_error"] = str(e)
        return None

def evaluate_day_formula(formula, col_values):
    import re
    match = re.match(r'^=([A-Z]+)1\+(\d+)$', formula)
    if match:
        ref_col_letter = match.group(1)
        offset = int(match.group(2))
        ref_col_num = 0
        for c in ref_col_letter:
            ref_col_num = ref_col_num * 26 + (ord(c) - ord('A') + 1)
        if ref_col_num in col_values:
            return col_values[ref_col_num] + offset
    return None

def import_excel_to_mongodb(filepath):
    db = get_db()
    if db is None:
        return False, "MongoDB nao disponivel"
    
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=False)
    xl = pd.ExcelFile(filepath)
    
    month_map = {
        "Janeiro": "Janeiro", "Fevereiro": "Fevereiro", "Marco": "Marco",
        "Março": "Marco", "Abril": "Abril", "Maio": "Maio",
        "Junho": "Junho", "Julho": "Julho", "Agosto": "Agosto",
        "Setembro": "Setembro", "Outubro": "Outubro", "Novembro": "Novembro",
        "Dezembro": "Dezembro"
    }
    
    db.vendas.drop()
    db.processed_days.drop()
    
    count = 0
    for sheet_name in xl.sheet_names:
        display_name = sheet_name
        for pt, clean in month_map.items():
            if pt in sheet_name:
                display_name = clean
                break
        
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        day_col_mapping = {}
        col_values = {}
        
        for col_idx in range(2, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if isinstance(val, (int, float)):
                day_num = int(float(val))
                if 1 <= day_num <= 31:
                    day_col_mapping[col_idx] = day_num
                    col_values[col_idx] = day_num
            elif isinstance(val, str) and val.startswith('='):
                calculated = evaluate_day_formula(val, col_values)
                if calculated and 1 <= calculated <= 31:
                    day_col_mapping[col_idx] = calculated
                    col_values[col_idx] = calculated
        
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        
        for row_idx in range(2, df.shape[0]):
            vendedor = str(df.iloc[row_idx, 0]).strip() if pd.notna(df.iloc[row_idx, 0]) else ""
            if not vendedor or vendedor == "nan":
                continue
            if "acumulado" in vendedor.lower() or "total" in vendedor.lower() or vendedor == "0":
                continue
            
            for col_idx, day_num in day_col_mapping.items():
                pandas_col = col_idx - 1
                val_sales = df.iloc[row_idx, pandas_col] if pandas_col < df.shape[1] else None
                try:
                    sales = float(val_sales) if pd.notna(val_sales) else 0.0
                except (ValueError, TypeError):
                    sales = 0.0
                
                client_col = pandas_col + 1
                if client_col < df.shape[1]:
                    val_clients = df.iloc[row_idx, client_col]
                    try:
                        clients = int(float(val_clients)) if pd.notna(val_clients) else 0
                    except (ValueError, TypeError):
                        clients = 0
                else:
                    clients = 0
                
                if sales != 0 or clients != 0:
                    db.vendas.insert_one({
                        "month": display_name,
                        "vendor": vendedor,
                        "day": day_num,
                        "sales": sales,
                        "clients": clients
                    })
                    count += 1
    
    wb.close()
    
    db.vendas.create_index([("month", 1), ("vendor", 1), ("day", 1)], unique=True)
    
    return True, f"{count} registos importados"

def load_monthly_data():
    db = get_db()
    if db is None:
        return None
    
    all_months = {}
    
    try:
        results = list(db.vendas.find({}, {"_id": 0}))
    except Exception:
        return None
    
    for r in results:
        month = r["month"]
        vendor = r["vendor"]
        day = r["day"]
        sales = r["sales"]
        clients = r["clients"]
        
        if month not in all_months:
            all_months[month] = {
                "days": set(),
                "vendedores": {},
                "vendedores_order": []
            }
        
        all_months[month]["days"].add(day)
        
        if vendor not in all_months[month]["vendedores"]:
            all_months[month]["vendedores"][vendor] = {
                "daily_sales": {},
                "daily_clients": {},
                "days": set()
            }
            all_months[month]["vendedores_order"].append(vendor)
        
        all_months[month]["vendedores"][vendor]["daily_sales"][day] = sales
        all_months[month]["vendedores"][vendor]["daily_clients"][day] = clients
        all_months[month]["vendedores"][vendor]["days"].add(day)
    
    for month in all_months:
        all_months[month]["days"] = sorted(all_months[month]["days"])
        for vendor in all_months[month]["vendedores"]:
            all_months[month]["vendedores"][vendor]["days"] = sorted(
                all_months[month]["vendedores"][vendor]["days"]
            )
    
    return all_months

def load_daily_files():
    daily_files = {}
    DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    
    for f in glob.glob(os.path.join(DATA_DIR, "*.xlsx")):
        fname = os.path.basename(f)
        if "vendas_mensais" in fname.lower():
            continue
        try:
            date_str = fname.replace(".xlsx", "")
            date_obj = datetime.strptime(date_str, "%d-%m-%Y")
            if date_obj not in daily_files:
                daily_files[date_obj] = f
        except ValueError:
            pass
    
    return daily_files

def load_single_daily(filepath):
    try:
        df = pd.read_excel(filepath, header=None)
        results = []
        current_resp = None
        
        for _, row in df.iterrows():
            val0 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            val1 = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
            val2 = row.iloc[2] if len(row) > 2 and pd.notna(row.iloc[2]) else None
            
            if "Resp." in val0 or "resp" in val0.lower():
                current_resp = val0
                continue
            
            if val0 and val0 != "nan" and val0.strip() != "" and val2 is not None:
                try:
                    total_liq = float(val2)
                    resp_display = current_resp.replace("Resp. Cobranca: ", "") if current_resp else "N/A"
                    resp_display = resp_display.replace("Resp. Cobran\u00e7a: ", "") if current_resp else "N/A"
                    results.append({
                        "Responsavel": resp_display,
                        "Entidade": val0,
                        "Nome": val1,
                        "Total Liq": total_liq
                    })
                except (ValueError, TypeError):
                    pass
        
        if results:
            return pd.DataFrame(results)
    except Exception as e:
        st.error(f"Erro ao ler ficheiro diario: {e}")
    return None

def parse_daily_file(filepath):
    try:
        df = pd.read_excel(filepath, header=None)
        results = []
        current_resp = None
        
        for _, row in df.iterrows():
            val0 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            val1 = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
            val2 = row.iloc[2] if len(row) > 2 and pd.notna(row.iloc[2]) else None
            
            if "Resp." in val0 or "resp" in val0.lower():
                current_resp = val0
                import re
                match = re.search(r'\((\d+)\)', val0)
                continue
            
            if val0 and val0 != "nan" and val0.strip() != "" and val2 is not None:
                try:
                    total_liq = float(val2)
                    results.append({
                        "Responsavel": current_resp,
                        "Entidade": val0,
                        "Nome": val1,
                        "Total Liq": total_liq
                    })
                except (ValueError, TypeError):
                    pass
        
        resp_summary = {}
        for r in results:
            resp = r["Responsavel"]
            if resp not in resp_summary:
                resp_summary[resp] = {"total_vendas": 0, "clientes": 0}
            resp_summary[resp]["total_vendas"] += r["Total Liq"]
        
        for resp_key in resp_summary:
            import re
            match = re.search(r'\((\d+)\)', resp_key)
            if match:
                resp_summary[resp_key]["clientes"] = int(match.group(1))
        
        return resp_summary
    except Exception as e:
        st.error(f"Erro ao ler ficheiro diario: {e}")
    return None

def get_vendedor_code_from_resp(resp_text):
    import re
    match = re.search(r':\s*(.+?)\s*$', resp_text)
    if match:
        code = match.group(1).strip()
        code = re.sub(r'\s*\(\d+\)\s*$', '', code)
        if code:
            return code
    return None

def update_monthly_data(day_date, daily_summary):
    db = get_db()
    if db is None:
        return False, "MongoDB nao disponivel"
    
    day_num = day_date.day
    month_num = day_date.month
    month_names = {
        1: "Janeiro", 2: "Fevereiro", 3: "Marco", 4: "Abril",
        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
    }
    month_name = month_names.get(month_num)
    
    processed = db.processed_days.find_one({"month": month_name, "day": day_num})
    if processed:
        return False, f"O dia {day_num} de {month_name} ja foi processado anteriormente"
    
    updated_vendors = []
    for resp_text, data in daily_summary.items():
        vendor_code = get_vendedor_code_from_resp(resp_text)
        
        if not vendor_code:
            vendor_code = "Sem Vendedor"
        
        try:
            existing = db.vendas.find_one({
                "month": month_name,
                "vendor": vendor_code,
                "day": day_num
            })
            
            if existing:
                new_sales = existing["sales"] + data["total_vendas"]
                new_clients = existing["clients"] + data["clientes"]
                db.vendas.update_one(
                    {"month": month_name, "vendor": vendor_code, "day": day_num},
                    {"$set": {"sales": new_sales, "clients": new_clients}}
                )
            else:
                db.vendas.insert_one({
                    "month": month_name,
                    "vendor": vendor_code,
                    "day": day_num,
                    "sales": data["total_vendas"],
                    "clients": data["clientes"]
                })
            
            updated_vendors.append(f"{vendor_code}: {data['total_vendas']:.2f} EUR ({data['clientes']} clientes)")
        except Exception as e:
            return False, f"Erro ao atualizar {vendor_code}: {str(e)}"
    
    db.processed_days.insert_one({"month": month_name, "day": day_num})
    
    return True, updated_vendors

def get_vendedor_cor(vendedor):
    cores = {
        "Vendas MT": "#1f77b4", "MT": "#1f77b4",
        "Vendas RC": "#ff7f0e", "RC ": "#ff7f0e", "RC": "#ff7f0e",
        "Vendas PC": "#2ca02c", "PC ": "#2ca02c", "PC": "#2ca02c",
        "Vendas RP": "#d62728", "RP ": "#d62728", "RP": "#d62728",
        "TR": "#9467bd", "NR": "#8c564b", "MS": "#e377c2",
        "GC": "#7f7f7f", "DR": "#bcbd22", "LF": "#17becf",
        "LO": "#aec7e8", "JM": "#ffbb78",
        "Vendas MTLis": "#1b9e77", "Vendas EG": "#d95f02",
        "M10 Leiria": "#7570b3", "M11 Évora": "#e7298a",
        "Viana MT1": "#66c2a5", "Braga MT2": "#fc8d62",
        "Vila Real MT3": "#8da0cb", "Bragança MT4": "#e78ac3",
        "Porto MT5": "#a6d854"
    }
    for key, cor in cores.items():
        if key.strip().lower() in vendedor.strip().lower():
            return cor
    return "#333333"

db = get_db()

if db is None:
    mongo_uri = os.environ.get("MONGODB_URI", "")
    if mongo_uri:
        mongo_error = st.session_state.get("mongo_error", "Erro desconhecido")
        st.sidebar.error(f"MongoDB nao disponivel: {mongo_error}")
        st.sidebar.info("Verifique se o IP do Render esta na whitelist do MongoDB Atlas (Network Access -> 0.0.0.0/0)")
    else:
        st.sidebar.warning("MongoDB nao configurado. Configure MONGODB_URI nas variaveis de ambiente.")
    
    uploaded_excel = st.sidebar.file_uploader("Upload ficheiro Excel mensal para importar", type=["xlsx"])
    if uploaded_excel:
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", uploaded_excel.name)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, "wb") as f:
            f.write(uploaded_excel.getbuffer())
        
        success, msg = import_excel_to_mongodb(filepath)
        if success:
            st.sidebar.success(msg)
            os.remove(filepath)
            st.rerun()
        else:
            st.sidebar.error(msg)
    
    st.stop()

monthly_data = load_monthly_data()
daily_files = load_daily_files()

st.sidebar.markdown("## Mapa de Vendas 2026")
st.sidebar.markdown("---")

if monthly_data:
    months_order = ["Janeiro", "Fevereiro", "Marco", "Abril", "Maio", "Junho",
                    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    available_months = [m for m in months_order if m in monthly_data]
    
    selected_month = st.sidebar.selectbox("Mes", available_months, index=len(available_months)-1 if available_months else 0)
    
    st.sidebar.markdown("### Filtro por Periodo")
    month_data = monthly_data.get(selected_month, None)
    
    if month_data:
        days = month_data["days"]
        col_a, col_b = st.sidebar.columns(2)
        with col_a:
            start_day = st.selectbox("Dia inicio", days, index=0)
        with col_b:
            end_day = st.selectbox("Dia fim", days, index=len(days)-1)
        
        filtered_days = [d for d in days if start_day <= d <= end_day]
        
        vendedores_disponiveis = [v for v in month_data["vendedores_order"] 
                                  if any(month_data["vendedores"][v]["daily_sales"].get(d, 0) != 0 for d in filtered_days)]
        
        selected_vendedores = st.sidebar.multiselect(
            "Vendedores",
            vendedores_disponiveis,
            default=vendedores_disponiveis
        )
else:
    st.sidebar.warning("Nenhum dado encontrado no MongoDB.")
    st.sidebar.info("Faça upload do ficheiro Excel mensal para importar os dados.")
    
    uploaded_excel = st.sidebar.file_uploader("Upload ficheiro Excel mensal", type=["xlsx"])
    if uploaded_excel:
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", uploaded_excel.name)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, "wb") as f:
            f.write(uploaded_excel.getbuffer())
        
        success, msg = import_excel_to_mongodb(filepath)
        if success:
            st.sidebar.success(msg)
            os.remove(filepath)
            st.rerun()
        else:
            st.sidebar.error(msg)
    st.stop()

tab1, tab2 = st.tabs(["Dashboard", "Analise Diaria"])

with tab1:
    st.markdown(f"# Mapa de Vendas - {selected_month} 2026")
    
    if month_data and filtered_days:
        total_vendas = 0
        vendedor_stats = {}
        
        for v in selected_vendedores:
            v_data = month_data["vendedores"].get(v, {})
            vendas_v = sum(v_data.get("daily_sales", {}).get(d, 0) for d in filtered_days)
            clientes_v = sum(v_data.get("daily_clients", {}).get(d, 0) for d in filtered_days)
            dias_trabalho = sum(1 for d in filtered_days if v_data.get("daily_sales", {}).get(d, 0) != 0)
            media_v = vendas_v / dias_trabalho if dias_trabalho > 0 else 0
            
            total_vendas += vendas_v
            
            vendedor_stats[v] = {
                "vendas": vendas_v,
                "clientes": clientes_v,
                "dias_trabalho": dias_trabalho,
                "media_diaria": media_v
            }
        
        clientes_por_vendedor_dia = {}
        for v in selected_vendedores:
            clientes_por_vendedor_dia[v] = {
                d: month_data["vendedores"].get(v, {}).get("daily_clients", {}).get(d, 0)
                for d in filtered_days
            }
        
        media_diaria_geral = total_vendas / len(filtered_days) if filtered_days else 0
        vendedores_ativos = sum(1 for v in selected_vendedores if vendedor_stats[v]["dias_trabalho"] > 0)
        
        kpi1, kpi2, kpi3 = st.columns(3)
        with kpi1:
            st.metric("Total Vendas", f"{total_vendas:,.2f} EUR")
        with kpi2:
            st.metric("Media Diaria", f"{media_diaria_geral:,.2f} EUR")
        with kpi3:
            st.metric("Vendedores Ativos", f"{vendedores_ativos}")
        
        st.markdown("---")
        
        col_chart1, col_chart2 = st.columns(2)
        
        with col_chart1:
            st.markdown("### Evolucao Diaria por Vendedor")
            fig_evol = go.Figure()
            for v in selected_vendedores:
                v_data = month_data["vendedores"].get(v, {})
                sales = [v_data.get("daily_sales", {}).get(d, 0) for d in filtered_days]
                fig_evol.add_trace(go.Scatter(
                    x=filtered_days, y=sales,
                    mode="lines+markers",
                    name=v.strip(),
                    line=dict(color=get_vendedor_cor(v), width=2),
                    marker=dict(size=5)
                ))
            fig_evol.update_layout(
                xaxis_title="Dia",
                yaxis_title="Vendas (EUR)",
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                height=400,
                margin=dict(l=40, r=20, t=40, b=40)
            )
            st.plotly_chart(fig_evol, use_container_width=True)
        
        with col_chart2:
            st.markdown("### Clientes por Dia (por Vendedor)")
            fig_clients = go.Figure()
            for v in selected_vendedores:
                if clientes_por_vendedor_dia.get(v):
                    fig_clients.add_trace(go.Bar(
                        x=list(filtered_days),
                        y=list(clientes_por_vendedor_dia[v].values()),
                        name=v.strip(),
                        marker_color=get_vendedor_cor(v)
                    ))
            fig_clients.update_layout(
                barmode="stack",
                xaxis_title="Dia",
                yaxis_title="Nº Clientes",
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                height=400,
                margin=dict(l=40, r=20, t=40, b=40)
            )
            st.plotly_chart(fig_clients, use_container_width=True)
        
        st.markdown("---")
        
        col_chart3, col_chart4 = st.columns(2)
        
        with col_chart3:
            st.markdown("### Acumulado por Vendedor")
            acum_data = {v: vendedor_stats[v]["vendas"] for v in selected_vendedores if vendedor_stats[v]["vendas"] > 0}
            if acum_data:
                fig_acum = go.Figure(data=[go.Pie(
                    labels=list(acum_data.keys()),
                    values=list(acum_data.values()),
                    hole=0.3,
                    marker=dict(colors=[get_vendedor_cor(v) for v in acum_data.keys()])
                )])
                fig_acum.update_layout(
                    height=400,
                    margin=dict(l=40, r=20, t=40, b=40)
                )
                st.plotly_chart(fig_acum, use_container_width=True)
        
        with col_chart4:
            st.markdown("### Media Diaria por Vendedor")
            media_data = {v.strip(): vendedor_stats[v]["media_diaria"] for v in selected_vendedores if vendedor_stats[v]["dias_trabalho"] > 0}
            if media_data:
                sorted_media = dict(sorted(media_data.items(), key=lambda x: x[1], reverse=True))
                fig_media = go.Figure(data=[go.Bar(
                    x=list(sorted_media.values()),
                    y=list(sorted_media.keys()),
                    orientation="h",
                    marker_color=[get_vendedor_cor(v) for v in sorted_media.keys()]
                )])
                fig_media.update_layout(
                    xaxis_title="Media Diaria (EUR)",
                    height=400,
                    margin=dict(l=40, r=20, t=40, b=40)
                )
                st.plotly_chart(fig_media, use_container_width=True)
        
        st.markdown("---")
        
        st.markdown("### Vendas por Semana de Trabalho")
        from datetime import date, timedelta
        year = 2026
        month_map = {"Janeiro": 1, "Fevereiro": 2, "Marco": 3, "Março": 3, "Abril": 4, "Maio": 5, "Junho": 6, "Julho": 7, "Agosto": 8, "Setembro": 9, "Outubro": 10, "Novembro": 11, "Dezembro": 12}
        month_num = month_map.get(selected_month, 6)
        
        weeks = {}
        for d in filtered_days:
            try:
                dt = date(year, month_num, d)
                weekday = dt.weekday()
                if weekday < 5:
                    week_start = dt - timedelta(days=weekday)
                    week_label = f"Sem {week_start.day}/{week_start.month}"
                    if week_label not in weeks:
                        weeks[week_label] = []
                    weeks[week_label].append(d)
            except ValueError:
                pass
        
        if weeks:
            week_names_sorted = sorted(weeks.keys(), key=lambda x: int(x.split()[1].split('/')[0]))
            
            week_chart_data = {}
            for v in selected_vendedores:
                v_data = month_data["vendedores"].get(v, {})
                week_chart_data[v] = {}
                for wk_label, wk_days in weeks.items():
                    week_chart_data[v][wk_label] = sum(v_data.get("daily_sales", {}).get(d, 0) for d in wk_days)
            
            fig_week = go.Figure()
            for v in selected_vendedores:
                fig_week.add_trace(go.Bar(
                    x=week_names_sorted,
                    y=[week_chart_data[v].get(wk, 0) for wk in week_names_sorted],
                    name=v.strip(),
                    marker_color=get_vendedor_cor(v)
                ))
            fig_week.update_layout(
                barmode="group",
                xaxis_title="Semana",
                yaxis_title="Vendas (EUR)",
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                height=400,
                margin=dict(l=40, r=20, t=40, b=40)
            )
            st.plotly_chart(fig_week, use_container_width=True)
            
            st.markdown("### Tabela de Vendas por Semana")
            week_table_data = []
            for v in selected_vendedores:
                row = {"Vendedor": v.strip()}
                for wk in week_names_sorted:
                    row[wk] = week_chart_data[v].get(wk, 0)
                row["Total"] = sum(week_chart_data[v].values())
                week_table_data.append(row)
            
            total_week_row = {"Vendedor": "TOTAL"}
            for wk in week_names_sorted:
                total_week_row[wk] = sum(week_chart_data[v].get(wk, 0) for v in selected_vendedores)
            total_week_row["Total"] = total_vendas
            week_table_data.append(total_week_row)
            
            df_week_table = pd.DataFrame(week_table_data)
            st.dataframe(df_week_table, use_container_width=True, hide_index=True)
        
        st.markdown("---")
        
        st.markdown("### Tabela de Vendas por Vendedor e Dia")
        table_data = []
        for v in selected_vendedores:
            v_data = month_data["vendedores"].get(v, {})
            row = {"Vendedor": v.strip()}
            for d in filtered_days:
                row[f"Dia {d}"] = v_data.get("daily_sales", {}).get(d, 0)
            row["Acumulado"] = vendedor_stats[v]["vendas"]
            row["Media/Dia"] = round(vendedor_stats[v]["media_diaria"], 2)
            table_data.append(row)
        
        df_table = pd.DataFrame(table_data)
        
        total_row = {"Vendedor": "TOTAL"}
        for d in filtered_days:
            total_row[f"Dia {d}"] = sum(
                month_data["vendedores"].get(v, {}).get("daily_sales", {}).get(d, 0)
                for v in selected_vendedores
            )
        total_row["Acumulado"] = total_vendas
        total_row["Media/Dia"] = round(media_diaria_geral, 2)
        df_table = pd.concat([df_table, pd.DataFrame([total_row])], ignore_index=True)
        
        st.dataframe(df_table, use_container_width=True, hide_index=True)
        
        with st.expander("Comparacao Mes a Mes"):
            st.markdown("### Comparacao de Vendas entre Meses")
            months_to_compare = st.multiselect(
                "Selecionar meses para comparar",
                available_months,
                default=[selected_month] if selected_month in available_months else available_months[:2]
            )
            
            if months_to_compare:
                comparison_data = []
                for m in months_to_compare:
                    m_data = monthly_data.get(m, {})
                    if m_data:
                        total_m = sum(
                            m_data.get("vendedores", {}).get(v, {}).get("daily_sales", {}).get(d, 0)
                            for v in m_data.get("vendedores_order", [])
                            for d in m_data.get("days", [])
                        )
                        clientes_m = sum(
                            m_data.get("vendedores", {}).get(v, {}).get("daily_clients", {}).get(d, 0)
                            for v in m_data.get("vendedores_order", [])
                            for d in m_data.get("days", [])
                        )
                        comparison_data.append({
                            "Mes": m,
                            "Total Vendas": total_m,
                            "Total Clientes": clientes_m
                        })
                
                df_comp = pd.DataFrame(comparison_data)
                
                fig_comp = make_subplots(rows=1, cols=2, subplot_titles=("Vendas Totais", "Total Clientes"))
                fig_comp.add_trace(go.Bar(
                    x=df_comp["Mes"], y=df_comp["Total Vendas"],
                    marker_color="#1f77b4", name="Vendas"
                ), row=1, col=1)
                fig_comp.add_trace(go.Bar(
                    x=df_comp["Mes"], y=df_comp["Total Clientes"],
                    marker_color="#ff7f0e", name="Clientes"
                ), row=1, col=2)
                fig_comp.update_layout(height=400, showlegend=False)
                st.plotly_chart(fig_comp, use_container_width=True)
                
                st.dataframe(df_comp, use_container_width=True, hide_index=True)

with tab2:
    st.markdown("# Analise Diaria")
    st.markdown("### Carregar ficheiro de vendas do dia")
    
    uploaded_daily = st.file_uploader("Carregar ficheiro Excel do dia (ex: 16-06-2026.xlsx)", type=["xlsx"], key="daily_upload")
    
    if uploaded_daily:
        try:
            date_str = uploaded_daily.name.replace(".xlsx", "")
            day_date = datetime.strptime(date_str, "%d-%m-%Y")
        except ValueError:
            st.error("Nome do ficheiro invalido. Use o formato DD-MM-AAAA.xlsx")
            st.stop()
        
        DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
        os.makedirs(DATA_DIR, exist_ok=True)
        temp_path = os.path.join(DATA_DIR, uploaded_daily.name)
        with open(temp_path, "wb") as f:
            f.write(uploaded_daily.getbuffer())
        
        daily_summary = parse_daily_file(temp_path)
        
        df_daily = load_single_daily(temp_path)
        
        if df_daily is not None:
            st.success(f"Ficheiro {uploaded_daily.name} carregado com sucesso!")
            
            total_dia = df_daily["Total Liq"].sum()
            clientes_dia = len(df_daily)
            responsaveis = df_daily["Responsavel"].unique()
            
            k1, k2, k3 = st.columns(3)
            with k1:
                st.metric("Total Vendas do Dia", f"{total_dia:,.2f} EUR")
            with k2:
                st.metric("N Clientes", f"{clientes_dia}")
            with k3:
                st.metric("Responsaveis", f"{len(responsaveis)}")
            
            st.markdown("---")
            
            st.markdown("### Vendas por Responsavel")
            resumo_resp = df_daily.groupby("Responsavel").agg(
                Total=("Total Liq", "sum"),
                Clientes=("Total Liq", "count")
            ).reset_index().sort_values("Total", ascending=False)
            
            fig_resp = go.Figure(data=[go.Bar(
                x=resumo_resp["Responsavel"],
                y=resumo_resp["Total"],
                text=resumo_resp["Total"].apply(lambda x: f"{x:,.2f}"),
                textposition="outside",
                marker_color="#1f77b4"
            )])
            fig_resp.update_layout(
                xaxis_title="Responsavel",
                yaxis_title="Total (EUR)",
                height=350
            )
            st.plotly_chart(fig_resp, use_container_width=True)
            
            st.markdown("---")
            st.markdown("### Atualizar Dados Mensais")
            
            if daily_summary:
                st.info(f"Datos extraidos do ficheiro diario:")
                for resp, data in daily_summary.items():
                    vendor_code = get_vendedor_code_from_resp(resp)
                    label = vendor_code if vendor_code else "Sem Vendedor"
                    st.write(f"- **{label}**: {data['total_vendas']:,.2f} EUR ({data['clientes']} clientes)")
                
                reprocess = st.checkbox("Reprocessar mesmo que o dia ja tenha sido processado")
                
                if st.button("Atualizar Dados Mensais", type="primary", use_container_width=True):
                    if reprocess:
                        db = get_db()
                        if db is not None:
                            month_names_map = {
                                1: "Janeiro", 2: "Fevereiro", 3: "Marco", 4: "Abril",
                                5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
                                9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
                            }
                            month_name_rp = month_names_map.get(day_date.month)
                            db.vendas.delete_many({"month": month_name_rp, "day": day_date.day})
                            db.processed_days.delete_many({"month": month_name_rp, "day": day_date.day})
                    success, message = update_monthly_data(day_date, daily_summary)
                    if success:
                        st.success("Dados mensais atualizados com sucesso!")
                        st.write("Vendedores atualizados:")
                        for v in message:
                            st.write(f"  - {v}")
                        
                        st.rerun()
                    else:
                        st.error(f"Erro ao atualizar: {message}")
            else:
                st.warning("Nao foi possivel extrair dados do ficheiro diario.")
            
            os.remove(temp_path)
        else:
            st.error("Nao foi possivel ler o ficheiro. Verifique o formato.")
    else:
        st.info("Ficheiros diarios disponiveis no sistema:")
        if daily_files:
            for date_obj in sorted(daily_files.keys(), reverse=True)[:10]:
                st.write(f"  {date_obj.strftime('%d-%m-%Y')}")
        else:
            st.write("  Nenhum ficheiro diario encontrado.")
        
        st.markdown("---")
        st.markdown("""
        **Instrucoes:**
        1. Faça upload do ficheiro Excel do dia (ex: `16-06-2026.xlsx`)
        2. Os dados serao guardados diretamente no MongoDB
        3. O ficheiro deve ter a estrutura: Entidade | Nome | Total Liq.
        4. Os dados persistem mesmo quando a app reinicia
        """)

st.sidebar.markdown("---")
st.sidebar.markdown("### Como atualizar dados")
st.sidebar.info("""
1. Na aba 'Analise Diaria', faca upload do ficheiro diario
2. Clique em 'Atualizar Dados Mensais'
3. Os dados sao guardados no MongoDB automaticamente
4. Nao e necessario ficheiro Excel local
""")
st.sidebar.markdown("---")
st.sidebar.caption("Dashboard de Vendas 2026")
